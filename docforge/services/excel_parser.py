"""
Parse .xlsx files into normalized records.

Header normalization maps common column variants ("Tenant Name", "name",
"Rent (€)") to the canonical field keys used by template packs.
"""
import re
from pathlib import Path

from openpyxl import load_workbook


_HEADER_ALIASES = {
    'tenant_name': {'tenant name', 'name', 'tenant', 'full name'},
    'address': {'address', 'property address', 'street'},
    'postal_code': {'postal code', 'postcode', 'zip', 'zip code'},
    'rent_amount': {'rent', 'rent amount', 'monthly rent', 'rent (eur)', 'rent (€)'},
    'arrears_amount': {'arrears', 'arrears amount', 'amount due', 'outstanding'},
    'email': {'email', 'e-mail'},
    'phone': {'phone', 'telephone', 'mobile'},
}


def _normalize_header(raw: str) -> str:
    if raw is None:
        return ''
    text = re.sub(r'\s+', ' ', str(raw)).strip().lower()
    for canonical, aliases in _HEADER_ALIASES.items():
        if text == canonical or text in aliases:
            return canonical
    return text.replace(' ', '_')


def parse_xlsx(path: str | Path) -> list[dict]:
    wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return []
    headers = [_normalize_header(h) for h in header_row]
    out: list[dict] = []
    for raw in rows_iter:
        if raw is None or all(v is None or str(v).strip() == '' for v in raw):
            continue
        record: dict = {}
        for key, val in zip(headers, raw):
            if not key:
                continue
            if val is None:
                record[key] = ''
            else:
                record[key] = str(val).strip()
        out.append(record)
    return out
